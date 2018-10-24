# -*- coding: utf-8 -*-
"""
File Name: read_mean.py
Author: GSS
Mail: gao.hillhill@gmail.com
Description: 
Created Time: 3/9/2016 7:12:33 PM
Last modified: Sun Oct 21 00:32:50 2018
"""

#defaut setting for scientific caculation
#import numpy
#import scipy
#from numpy import *
#import numpy as np
#import scipy as sp
#import pylab as pl

import openpyxl as px
import numpy as np
import statsmodels.api as sm
import matplotlib.pyplot as plt
import sys
import os

def readxlsx(path):
    #vW = px.load_workbook(path)
    vW = px.load_workbook(path)
    vp = vW.get_sheet_by_name(name = 'sheet1')
    
    va=[]
    vi = 0
    for row in vp.iter_rows():
        if vi > 0:
            for k in row:
                va.append(float(k.internal_value))
        vi = vi + 1
    
    # convert list a to matrix (for example 16*6)
    vaa= np.resize(va, [vi, 1 ])
    adc = vaa[:,0]
    return adc

def oneplot(plt, data, clor = "C1", plt_smps = 200, label="20fC"):
    ppeak = np.max(data)
    ppos = np.where( data == ppeak) [0][0]
    subdata = data[ppos-1100:ppos+1100]
    ped = np.mean(data[ppos+5000:ppos+10000])
    subdata = np.array(subdata) - int(ped)
    amp = np.max(subdata)

#    ppos2f = np.where( subdata > (amp/2.0)) [0][0]
#    ppos2e = np.where( subdata > (amp/2.0)) [0][-1]
#    hwidth = ppos2e - ppos2f + 1
  
    m = 30
    for i in range(len(subdata)):
        if ((subdata[i+1]-subdata[i])>m) and ((subdata[i+2]-subdata[i+1])>m) and \
            ((subdata[i+3]-subdata[i+2])>m) and (subdata[i] >= 0 ):
            pstart = i
            break
    n = -10
    for i in range(len(subdata)-10):
        if ((subdata[i+1]-subdata[i])<n) and ((subdata[i+2]-subdata[i+1])<n) and \
            ((subdata[i+3]-subdata[i+2])<n) and (subdata[i] >= 0 ):
            pend = i
 
    parea = np.sum(subdata[pstart:pend])
    width = pend-pstart

    subdata = subdata[pstart-50:]
    subdata = subdata[0:plt_smps]
    amppos = np.where( subdata == amp) [0][0]
    x = range(len(subdata))
    plt.scatter(x, subdata, c = clor, marker = '.' )
    plt.plot(x, subdata, c = clor, label = label)
#    plt.text(amppos+1, amp, "%d"%amp, color = clor)
    print amp, width, parea
    return amp, width, parea

path = "/Users/shanshangao/Google_Drive_BNL/tmp/ledge/5mvperus/"
for root, dirs, files in os.walk(path):
    break

mv10 = [ [5, "001"], [10, "002"], [20,"003"],   [50,"004"],   [100,"005"], 
         [200,"006"],  [300,"007"], [400, "008"], [500, "009"], [750,"010"], 
         [1000,"011"], [2000,"012"] 
        ]
#         [1000,"018"], [1500,"017"], [2000,"019"] ]

#plt_smps = 100
#fig = plt.figure(figsize=(12,8))
#for i in range(9):
#    for f in files:
#        if (f.find(mv10[i][1] + ".xlsx")>=0 ) and (f.find("per")>=0 ) :
#            fp = path + f
#            data = readxlsx(fp)
#            fC = int( mv10[i][0]*1.203 )
#            b = oneplot(plt, data, clor = "C%d"%((i+1)%9), plt_smps = plt_smps, label ="%d fC"%fC  )
#
#plt.title( "12 fC/us", fontsize = 16 )
#plt.ylabel("( ADC output - Pedestal ) / bin", fontsize = 16 )
#plt.ylim([-100,2000])
#plt.xlabel("ADC ticks", fontsize = 16 )
#plt.xlim([0,plt_smps])
#plt.legend(loc='best', fontsize = 16)
#plt.tick_params(labelsize=16)
#plt.legend(loc=1, fontsize = 16)
#plt.grid()
#plt.tight_layout( rect=[0.05, 0.05, 0.95, 0.95])
#plt.savefig("./10mVperus_240fC.png")
#plt.close()


plt_smps = 1000
fig = plt.figure(figsize=(12,8))
paras = []
for i in range(len(mv10)):
    for f in files:
        if (f.find(mv10[i][1] + ".xlsx")>=0 ) and (f.find("per")>=0 ) :
            fp = path + f
            data = readxlsx(fp)
            fC = int( mv10[i][0]*1.203 )
            b = oneplot(plt, data, clor = "C%d"%((i+1)%9), plt_smps = plt_smps, label ="%d fC"%fC  )
            paras.append(b)

plt.title( "6 fC/us", fontsize = 16 )
plt.ylabel("( ADC output - Pedestal ) / bin", fontsize = 16 )
plt.ylim([-100,2000])
plt.xlabel("ADC ticks", fontsize = 16 )
plt.xlim([0,plt_smps])
plt.legend(loc='best', fontsize = 16)
plt.tick_params(labelsize=16)
plt.legend(loc=1, fontsize = 16)
plt.grid()
plt.tight_layout( rect=[0.05, 0.05, 0.95, 0.95])
plt.savefig("./5mVperus_2000mV.png")
plt.close()

amps = []
widths = []
areas = []
for p in paras:
    amps.append(p[0])
    widths.append(p[1])
    areas.append(p[2])
plen = len(paras) 

fcs = []
for tmp in mv10:
    fcs.append(tmp[0]*1.203)

fig = plt.figure(figsize=(12,8))
plt.scatter(fcs[0:plen], areas[0:plen], c = 'r', marker = 'x' )

plin = plen -2
cresults = sm.OLS(areas[0:plin],sm.add_constant(fcs[0:plin])).fit()
cslope = cresults.params[1]
cconstant = cresults.params[0]
label = "Y = (%d) * X + (%d)"%(cslope, cconstant)

#plt.plot(fcs[0:plen], areas, c = 'b', label = label)
plt.plot(fcs[0:plen], np.array(fcs[0:plen]) * cslope + cconstant, c = 'b', label = label)
plt.title( "6 fC/us", fontsize = 16 )
plt.ylabel(" (ADC output - Pedestal) * Ticks", fontsize = 16 )
plt.ylim([0,400000])
plt.xlabel("Injected charge / fC", fontsize = 16 )
plt.xlim([0, 2500])
plt.legend(loc='best', fontsize = 16)
plt.tick_params(labelsize=16)
plt.legend(loc="best", fontsize = 16)
plt.grid()
plt.tight_layout( rect=[0.05, 0.05, 0.95, 0.95])
plt.savefig("./5mVperus_area.png")
plt.close()


#fig = plt.figure(figsize=(12,8))
#plt.scatter(fcs[0:plen], amps, c = 'b', marker = '.' )
#plt.plot(fcs[0:plen], amps, c = 'b', label = "Amplitude")
#plt.title( "12 fC/us", fontsize = 16 )
#plt.ylabel("(ADC output - Pedestal) / bin", fontsize = 16 )
#plt.ylim([0,2000])
#plt.xlabel("Injected charge / fC", fontsize = 16 )
#plt.xlim([0, max(fcs)])
#plt.legend(loc='best', fontsize = 16)
#plt.tick_params(labelsize=16)
#plt.legend(loc=1, fontsize = 16)
#plt.grid()
#plt.tight_layout( rect=[0.05, 0.05, 0.95, 0.95])
#plt.savefig("./10mVperus_amp.png")
#plt.close()

#fig = plt.figure(figsize=(12,8))
#plt.scatter(fcs[0:plen], widths, c = 'b', marker = '.' )
#plt.plot(fcs[0:plen], widths, c = 'b', label = "Amp")
#plt.show()
#plt.close()


